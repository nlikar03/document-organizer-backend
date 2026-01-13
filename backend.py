import os
import json
import zipfile
import io
import fitz
from typing import List, Dict, Any
from concurrent.futures import ThreadPoolExecutor, as_completed
import tempfile


from fastapi import FastAPI, UploadFile, File, HTTPException, Header, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from PIL import Image
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from openai import OpenAI

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from google.cloud import vision
from google.oauth2 import service_account

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:5173",
        "*"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = OpenAI()

# Initialize Google Vision with credentials from environment variable
GOOGLE_CREDENTIALS_JSON = os.getenv("GOOGLE_CREDENTIALS_JSON")
if GOOGLE_CREDENTIALS_JSON:
    credentials_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
    credentials = service_account.Credentials.from_service_account_info(credentials_dict)
    vision_client = vision.ImageAnnotatorClient(credentials=credentials)
else:
    vision_client = vision.ImageAnnotatorClient()

APP_PASSWORD = os.getenv("APP_PASSWORD", "default_password")


# MODELS
class AIRequest(BaseModel):
    text: str
    structure: List[Dict[str, Any]]


class AIBatchRequest(BaseModel):
    texts: List[str]
    fileNames: List[str]
    structure: List[Dict[str, Any]]


class ExcelRequest(BaseModel):
    results: List[Dict[str, Any]]
    structure: List[Dict[str, Any]]


async def verify_password(x_password: str = Header(None)):
    if x_password != APP_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    return True


# GOOGLE VISION OCR

def extract_text_with_vision(image_bytes: bytes) -> str:
    """
    OCR using Google Cloud Vision API - much faster than OpenAI.
    """
    try:
        image = vision.Image(content=image_bytes)
        response = vision_client.text_detection(image=image)
        
        if response.error.message:
            raise Exception(f"Vision API error: {response.error.message}")
        
        texts = response.text_annotations
        if texts:
            return texts[0].description
        return ""
    
    except Exception as e:
        return f"Vision OCR error: {str(e)}"


def extract_text_from_pdf_batch(pdf_data: bytes, max_pages: int = 5) -> str:
    """
    Extract text from PDF using Google Vision with parallel processing.
    Much faster than sequential page-by-page requests.
    """
    try:
        doc = fitz.open(stream=pdf_data, filetype="pdf")
        pages_to_process = min(len(doc), max_pages)
        
        page_images = []
        for page_num in range(pages_to_process):
            page = doc[page_num]
            pix = page.get_pixmap(dpi=150)  #dpi loh je nizji
            img_bytes = pix.tobytes("png")
            page_images.append((page_num, img_bytes))
        
        full_text = [""] * pages_to_process
        
        with ThreadPoolExecutor(max_workers=5) as executor:
            future_to_page = {
                executor.submit(extract_text_with_vision, img_bytes): page_num 
                for page_num, img_bytes in page_images
            }
            
            for future in as_completed(future_to_page):
                page_num = future_to_page[future]
                try:
                    text = future.result()
                    full_text[page_num] = text
                except Exception as e:
                    full_text[page_num] = f"Error on page {page_num + 1}: {str(e)}"
        
        doc.close()
        return "\n\n".join(full_text)
    
    except Exception as e:
        return f"PDF processing error: {str(e)}"


def extract_text_from_image(image_bytes: bytes) -> str:
    """
    OCR a single image using Google Vision.
    """
    return extract_text_with_vision(image_bytes)


def classify_single_document(text: str, structure: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Classify a single document synchronously.
    """
    simplified = [
        {"id": f["id"], "name": f["name"], "level": f["level"]}
        for f in structure
    ]

    prompt = f"""
You are a document organizer assistant.

DOCUMENT TEXT:
{text[:8000]}

FOLDER STRUCTURE:
{json.dumps(simplified, ensure_ascii=False)}

Analyze this document and return ONLY JSON with these fields:
{{
  "suggestedFolder": {{ "id": "folder_id", "name": "folder_name" }},
  "documentTitle": "brief description of what this document is about",
  "issuer": "who issued/created this document",
  "documentNumber": "document number if visible (or empty string)",
  "date": "document date in DD.MM.YYYY format if visible (or empty string)"
}}
"""

    try:
        response = client.chat.completions.create(
            model="gpt-5-mini",
            messages=[
                {"role": "system", "content": "You are a document classification assistant. Always respond with valid JSON."},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"}
        )

        result = json.loads(response.choices[0].message.content)
        folder_data = result.get("suggestedFolder", {})
        folder_id = folder_data.get("id", "0")

        return {
            "suggestedFolder": {
                "id": folder_id,
                "name": folder_data.get("name", "Unknown"),
                "fullPath": build_folder_path(folder_id, structure)
            },
            "documentTitle": result.get("documentTitle", ""),
            "issuer": result.get("issuer", ""),
            "documentNumber": result.get("documentNumber", ""),
            "date": result.get("date", "")
        }

    except Exception as e:
        return {
            "suggestedFolder": {
                "id": "0",
                "name": "Unknown",
                "fullPath": "0 PODATKI O POGODBI"
            },
            "documentTitle": "",
            "issuer": "",
            "documentNumber": "",
            "date": ""
        }


# PDF WATERMARK

def add_watermark_to_pdf(pdf_bytes: bytes, watermark_text: str) -> bytes:
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        if not reader.pages:
            return pdf_bytes

        first_page = reader.pages[0]
        width = float(first_page.mediabox.width)
        height = float(first_page.mediabox.height)

        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=(width, height))
        can.setFillColorRGB(1, 0, 0)
        can.setFont("Helvetica-Bold", 20)

        margin = 30
        can.drawRightString(width - margin, height - margin, watermark_text)
        can.save()

        packet.seek(0)
        watermark_pdf = PdfReader(packet)

        writer = PdfWriter()
        first_page.merge_page(watermark_pdf.pages[0])
        writer.add_page(first_page)

        for i in range(1, len(reader.pages)):
            writer.add_page(reader.pages[i])

        output = io.BytesIO()
        writer.write(output)
        output.seek(0)
        return output.read()

    except Exception:
        return pdf_bytes


# HELPERJI

def build_folder_path(folder_id: str, structure: List[Dict[str, Any]]) -> str:
    folder_map = {f["id"]: f["name"] for f in structure}
    parts = folder_id.split(".")
    path = []

    for i in range(len(parts)):
        pid = ".".join(parts[:i + 1])
        if pid in folder_map:
            path.append(folder_map[pid])

    return " / ".join(path)


# ROUTES

@app.post("/api/verify-password")
async def verify_password_endpoint(valid: bool = Depends(verify_password)):
    return {"valid": True}


@app.post("/api/ocr", dependencies=[Depends(verify_password)])
async def process_ocr(file: UploadFile = File(...)):
    contents = await file.read()
    name = file.filename.lower()

    if name.endswith(".pdf"):
        text = extract_text_from_pdf_batch(contents)
    elif name.endswith((".png", ".jpg", ".jpeg", ".bmp")):
        text = extract_text_from_image(contents)
    else:
        raise HTTPException(400, "Unsupported file format")

    return {"fileName": file.filename, "text": text}


@app.post("/api/classify", dependencies=[Depends(verify_password)])
async def classify_document(request: AIRequest):
    result = classify_single_document(request.text, request.structure)
    return result


@app.post("/api/classify-batch", dependencies=[Depends(verify_password)])
async def classify_batch(request: AIBatchRequest):
    """
    Classify multiple documents in parallel (max 5 concurrent).
    """
    try:
        results = []
        
        with ThreadPoolExecutor(max_workers=5) as executor:
            future_to_index = {
                executor.submit(classify_single_document, text, request.structure): idx
                for idx, text in enumerate(request.texts)
            }
            
            results = [None] * len(request.texts)
            for future in as_completed(future_to_index):
                idx = future_to_index[future]
                try:
                    result = future.result()
                    results[idx] = {
                        "fileName": request.fileNames[idx],
                        "classification": result
                    }
                except Exception as e:  #default exception
                    results[idx] = {
                        "fileName": request.fileNames[idx],
                        "classification": {
                            "suggestedFolder": {
                                "id": "0",
                                "name": "Unknown",
                                "fullPath": "0 PODATKI O POGODBI"
                            },
                            "documentTitle": "",
                            "issuer": "",
                            "documentNumber": "",
                            "date": "",
                            "error": str(e)
                        }
                    }
        
        return {"results": results}
        
    except Exception as e:
        raise HTTPException(500, str(e))
    
def iter_file(path, chunk_size=1024 * 1024):
    with open(path, "rb") as f:
        while chunk := f.read(chunk_size):
            yield chunk

@app.post("/api/generate-zip", dependencies=[Depends(verify_password)])
async def generate_zip(
    files: List[UploadFile] = File(...),
    metadata: UploadFile = File(...)
):
    try:
        meta = json.loads((await metadata.read()).decode())
        structure = meta["structure"]
        file_mapping = meta["fileMapping"]


        tmp = tempfile.NamedTemporaryFile(delete=False)
        zip_path = tmp.name
        tmp.close()

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            folder_paths = {}

            for folder in structure:
                parts = []
                ids = folder["id"].split(".")
                for i in range(len(ids)):
                    fid = ".".join(ids[:i + 1])
                    f = next((x for x in structure if x["id"] == fid), None)
                    if f:
                        parts.append(f["name"])
                path = os.path.join(*parts)
                folder_paths[folder["id"]] = path
                zipf.writestr(path + "/", "")

            for f in files:
                info = file_mapping.get(f.filename)
                if not info:
                    continue

                content = await f.read()

                if f.filename.lower().endswith(".pdf"):
                    content = add_watermark_to_pdf(content, info["docCode"])

                base, ext = os.path.splitext(f.filename)
                new_name = f"{info['fileNumber']:03d}_{base}{ext}"
                path = os.path.join(folder_paths[info["folderId"]], new_name)

                zipf.writestr(path, content)

        return StreamingResponse(
            iter_file(zip_path),
            media_type="application/zip",
            headers={
                "Content-Disposition": "attachment; filename=DZO_Dokumenti.zip"
            }
        )

    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/generate-excel", dependencies=[Depends(verify_password)])
async def generate_excel(request: ExcelRequest):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Seznam Dokumentov"
        
        headers = ["Zap. št.", "Ime dokazila oz. na kaj se dokazilo nanaša", 
                   "Izdajatelj", "Št. dokazila", "Datum", "Kategorija", "Koda dokumenta"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        sorted_results = sorted(request.results, key=lambda x: x['suggestedFolder']['id'])
        
        for res in sorted_results:
            ws.append([
                res['fileNumber'],
                res.get('documentTitle', ''),
                res.get('issuer', ''),
                res.get('documentNumber', ''),
                res.get('date', ''),
                build_folder_path(res['suggestedFolder']['id'], request.structure),
                res['docCode']
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=DZO_Dokumenti_Seznam.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(500, str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)